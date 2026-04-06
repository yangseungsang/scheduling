import csv
import io
from datetime import date, datetime, timedelta


HEADERS = ['날짜', '장절명']


def _block_to_row(b):
    return [
        b.get('date', ''),
        b.get('section_name', '') or b.get('task_title', ''),
    ]


def export_csv(enriched_blocks):
    """Generate CSV string with BOM for Korean Excel compatibility."""
    buf = io.StringIO()
    buf.write('\ufeff')
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for b in enriched_blocks:
        writer.writerow(_block_to_row(b))
    return buf.getvalue()


def export_xlsx(enriched_blocks, start_date, end_date):
    """Generate calendar-layout XLSX as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '스케줄'

    # Group blocks by date
    blocks_by_date = {}
    for b in enriched_blocks:
        blocks_by_date.setdefault(b.get('date', ''), []).append(b)

    # Build calendar weeks
    d_start = datetime.strptime(start_date, '%Y-%m-%d').date()
    d_end = datetime.strptime(end_date, '%Y-%m-%d').date()
    week_start = d_start - timedelta(days=d_start.weekday())
    week_end = d_end + timedelta(days=6 - d_end.weekday())

    day_names = ['월', '화', '수', '목', '금', '토', '일']

    # Styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    date_font = Font(bold=True, size=10)
    block_font = Font(size=9)
    today_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    weekend_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    top_align = Alignment(vertical='top', wrap_text=True)
    today = date.today()

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1, value=f'스케줄: {start_date} ~ {end_date}')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 22

    row = 3
    current = week_start
    while current <= week_end:
        # Day-name header row
        for i in range(7):
            cell = ws.cell(row=row, column=i + 1, value=day_names[i])
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        # Date row
        for i in range(7):
            day = current + timedelta(days=i)
            label = day.strftime('%m/%d') if d_start <= day <= d_end else ''
            cell = ws.cell(row=row, column=i + 1, value=label)
            cell.font = date_font
            cell.alignment = center_align
            cell.border = thin_border
            if day == today:
                cell.fill = today_fill
            elif day.weekday() >= 5:
                cell.fill = weekend_fill
        row += 1

        # Max blocks this week
        max_blocks = 0
        for i in range(7):
            day = current + timedelta(days=i)
            max_blocks = max(max_blocks, len(blocks_by_date.get(day.isoformat(), [])))
        content_rows = max(max_blocks, 1)

        # Fill block content rows
        for r_offset in range(content_rows):
            for i in range(7):
                day = current + timedelta(days=i)
                day_blocks = blocks_by_date.get(day.isoformat(), [])

                cell = ws.cell(row=row + r_offset, column=i + 1)
                cell.border = thin_border
                cell.alignment = top_align

                if day.weekday() >= 5:
                    cell.fill = weekend_fill
                if day == today:
                    cell.fill = today_fill

                if r_offset < len(day_blocks):
                    b = day_blocks[r_offset]
                    section = b.get('section_name', '') or b.get('task_title', '')
                    cell.value = section
                    cell.font = block_font
                    # Tint cell with block color
                    color_hex = (b.get('color') or '#FFFFFF').lstrip('#')
                    if len(color_hex) == 6:
                        r_c = int(int(color_hex[0:2], 16) * 0.3 + 255 * 0.7)
                        g_c = int(int(color_hex[2:4], 16) * 0.3 + 255 * 0.7)
                        b_c = int(int(color_hex[4:6], 16) * 0.3 + 255 * 0.7)
                        light = f'{r_c:02X}{g_c:02X}{b_c:02X}'
                        cell.fill = PatternFill(start_color=light, end_color=light, fill_type='solid')

        for r_offset in range(content_rows):
            ws.row_dimensions[row + r_offset].height = 60

        row += content_rows
        row += 1  # blank separator
        current += timedelta(days=7)

    # Data sheet
    ws2 = wb.create_sheet(title='데이터')
    ws2.append(HEADERS)
    for b in enriched_blocks:
        ws2.append(_block_to_row(b))
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ''
            max_len = max(max_len, len(val))
        ws2.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
