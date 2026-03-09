import calendar
import csv
import io
from datetime import date, datetime, timedelta

from flask import Blueprint, request, jsonify, render_template, Response

from app.repositories import (
    category_repo,
    schedule_repo,
    settings_repo,
    task_repo,
    user_repo,
)

schedule_bp = Blueprint('schedule', __name__, url_prefix='/schedule')


def _parse_date(date_str):
    """Parse a date string (YYYY-MM-DD) or return today."""
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    return date.today()


def _enrich_blocks(blocks, users_map, tasks_map, categories_map, color_by):
    """Add task title, assignee name/color, category name/color to blocks."""
    enriched = []
    for b in blocks:
        block = dict(b)
        task = tasks_map.get(b.get('task_id'))
        assignee = users_map.get(b.get('assignee_id'))
        category = None
        if task:
            category = categories_map.get(task.get('category_id'))

        block['task_title'] = task['title'] if task else '(삭제된 업무)'
        block['priority'] = task.get('priority', '') if task else ''
        block['assignee_name'] = assignee['name'] if assignee else '(미배정)'
        block['assignee_color'] = assignee['color'] if assignee else '#6c757d'
        block['category_name'] = category['name'] if category else ''
        block['category_color'] = category['color'] if category else '#6c757d'

        if color_by == 'category':
            block['color'] = block['category_color']
        else:
            block['color'] = block['assignee_color']

        enriched.append(block)
    return enriched


def _build_maps():
    """Build lookup maps for users, tasks, categories."""
    users = user_repo.get_all()
    tasks = task_repo.get_all()
    categories = category_repo.get_all()
    return (
        {u['id']: u for u in users},
        {t['id']: t for t in tasks},
        {c['id']: c for c in categories},
    )


def _get_queue_tasks(users_map, categories_map):
    """Get tasks with remaining unscheduled hours for the task queue sidebar."""
    tasks = task_repo.get_all()
    all_blocks = schedule_repo.get_all()

    # Sum up already-scheduled hours per task
    scheduled_hours = {}
    for b in all_blocks:
        tid = b['task_id']
        start_min = _time_to_minutes(b['start_time'])
        end_min = _time_to_minutes(b['end_time'])
        scheduled_hours[tid] = scheduled_hours.get(tid, 0) + (end_min - start_min) / 60.0

    queue = []
    for t in tasks:
        if t['status'] == 'completed' or t.get('remaining_hours', 0) <= 0:
            continue
        remaining_unscheduled = t.get('remaining_hours', 0) - scheduled_hours.get(t['id'], 0)
        if remaining_unscheduled <= 0:
            continue
        task = dict(t)
        task['remaining_unscheduled_hours'] = round(remaining_unscheduled, 2)
        assignee = users_map.get(t.get('assignee_id'))
        category = categories_map.get(t.get('category_id'))
        task['assignee_name'] = assignee['name'] if assignee else '(미배정)'
        task['assignee_color'] = assignee['color'] if assignee else '#6c757d'
        task['category_name'] = category['name'] if category else ''
        task['category_color'] = category['color'] if category else '#6c757d'
        queue.append(task)
    # Sort: deadline ascending, then priority
    priority_order = {'high': 0, 'medium': 1, 'low': 2}
    queue.sort(key=lambda t: (
        t.get('deadline') or '9999-12-31',
        priority_order.get(t.get('priority', 'low'), 2),
    ))
    return queue


def _generate_time_slots(settings):
    """Generate list of time slot strings from work_start to work_end."""
    interval = settings.get('grid_interval_minutes', 15)
    start = datetime.strptime(settings['work_start'], '%H:%M')
    end = datetime.strptime(settings['work_end'], '%H:%M')
    slots = []
    current = start
    while current < end:
        slots.append(current.strftime('%H:%M'))
        current += timedelta(minutes=interval)
    return slots


def _is_break_slot(time_str, settings):
    """Check if a time slot falls within lunch or break periods."""
    lunch_start = settings.get('lunch_start', '12:00')
    lunch_end = settings.get('lunch_end', '13:00')
    if lunch_start <= time_str < lunch_end:
        return True
    for brk in settings.get('breaks', []):
        if brk['start'] <= time_str < brk['end']:
            return True
    return False


# ---------------------------------------------------------------------------
# Break-aware end time adjustment
# ---------------------------------------------------------------------------

def _time_to_minutes(t):
    """Convert 'HH:MM' to total minutes."""
    parts = t.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def _minutes_to_time(m):
    """Convert total minutes to 'HH:MM'."""
    return f'{m // 60:02d}:{m % 60:02d}'


def _get_break_periods(settings):
    """Return sorted list of (start_min, end_min) for lunch + breaks."""
    periods = []
    lunch_s = settings.get('lunch_start', '12:00')
    lunch_e = settings.get('lunch_end', '13:00')
    periods.append((_time_to_minutes(lunch_s), _time_to_minutes(lunch_e)))
    for brk in settings.get('breaks', []):
        periods.append((_time_to_minutes(brk['start']), _time_to_minutes(brk['end'])))
    periods.sort()
    return periods


def _adjust_end_for_breaks(start_time, end_time, settings):
    """Adjust end_time so that actual work duration is preserved.

    For each break period that falls within [start_time, end_time],
    the end_time is pushed forward by the break's duration.
    This is applied iteratively since pushing end_time may expose
    additional breaks.
    """
    work_end = _time_to_minutes(settings.get('work_end', '18:00'))
    start_min = _time_to_minutes(start_time)
    end_min = _time_to_minutes(end_time)
    work_duration = end_min - start_min
    if work_duration <= 0:
        return end_time

    breaks = _get_break_periods(settings)

    # Walk forward from start_time, accumulating work minutes and
    # skipping break periods until we've placed enough work time.
    current = start_min
    remaining_work = work_duration

    while remaining_work > 0:
        # Find the next break that starts at or after 'current'
        # or overlaps with current position
        next_break = None
        for bs, be in breaks:
            if be <= current:
                continue  # break already passed
            if bs <= current:
                # We're inside a break, skip to end of it
                current = be
                continue
            next_break = (bs, be)
            break

        if next_break is None:
            # No more breaks ahead; place all remaining work
            current += remaining_work
            remaining_work = 0
        else:
            bs, be = next_break
            available = bs - current
            if available >= remaining_work:
                # Enough room before the next break
                current += remaining_work
                remaining_work = 0
            else:
                # Fill up to break, skip break, continue
                remaining_work -= available
                current = be  # jump past break

    # Cap at work_end
    if current > work_end:
        current = work_end

    return _minutes_to_time(current)


def _work_minutes_in_range(start_time, end_time, settings):
    """Calculate actual work minutes in a time range, excluding breaks."""
    breaks = _get_break_periods(settings)
    start_min = _time_to_minutes(start_time)
    end_min = _time_to_minutes(end_time)
    total = end_min - start_min
    for bs, be in breaks:
        overlap_start = max(start_min, bs)
        overlap_end = min(end_min, be)
        if overlap_start < overlap_end:
            total -= (overlap_end - overlap_start)
    return max(0, total)


# ---------------------------------------------------------------------------
# Overlap check
# ---------------------------------------------------------------------------

def _check_overlap(assignee_id, date_str, start_time, end_time, exclude_block_id=None):
    """Check if a block would overlap with existing blocks for the same assignee."""
    if not assignee_id:
        return None
    existing = schedule_repo.get_by_date(date_str)
    s1 = _time_to_minutes(start_time)
    e1 = _time_to_minutes(end_time)
    for b in existing:
        if b['assignee_id'] != assignee_id:
            continue
        if exclude_block_id and b['id'] == exclude_block_id:
            continue
        s2 = _time_to_minutes(b['start_time'])
        e2 = _time_to_minutes(b['end_time'])
        if s1 < e2 and s2 < e1:
            return b
    return None


# ---------------------------------------------------------------------------
# Overlap layout computation
# ---------------------------------------------------------------------------

def _compute_overlap_layout(blocks):
    """Compute left/width for overlapping blocks so they sit side by side.

    Each block gets ``col_index`` and ``col_total`` attributes added.
    Uses a greedy column-packing algorithm:
    1. Sort blocks by start_time, then by end_time desc.
    2. Maintain a list of columns; each column tracks its latest end_time.
    3. Assign each block to the first column whose end <= block.start,
       or create a new column.
    4. After all blocks are assigned, walk through each group of
       overlapping blocks to set col_total to the max columns in that group.
    """
    if not blocks:
        return blocks

    sorted_blocks = sorted(
        blocks,
        key=lambda b: (_time_to_minutes(b['start_time']),
                       -_time_to_minutes(b['end_time'])),
    )

    columns = []  # list of (end_min, [block_indices])
    block_col = {}  # block index -> column index

    for i, b in enumerate(sorted_blocks):
        s = _time_to_minutes(b['start_time'])
        placed = False
        for ci, (col_end, _) in enumerate(columns):
            if col_end <= s:
                columns[ci] = (_time_to_minutes(b['end_time']), columns[ci][1] + [i])
                block_col[i] = ci
                placed = True
                break
        if not placed:
            block_col[i] = len(columns)
            columns.append((_time_to_minutes(b['end_time']), [i]))

    # For each block, find the max number of columns among all blocks
    # that overlap with it (its overlap group).
    for i, b in enumerate(sorted_blocks):
        s = _time_to_minutes(b['start_time'])
        e = _time_to_minutes(b['end_time'])
        max_col = block_col[i] + 1
        for j, b2 in enumerate(sorted_blocks):
            if i == j:
                continue
            s2 = _time_to_minutes(b2['start_time'])
            e2 = _time_to_minutes(b2['end_time'])
            if s < e2 and s2 < e:  # overlaps
                c = block_col[j] + 1
                if c > max_col:
                    max_col = c
        b['col_index'] = block_col[i]
        b['col_total'] = max_col

    return sorted_blocks


# ---------------------------------------------------------------------------
# Template rendering routes
# ---------------------------------------------------------------------------

@schedule_bp.route('/')
def day_view():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    blocks = schedule_repo.get_by_date(current_date.isoformat())
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )
    enriched = _compute_overlap_layout(enriched)

    time_slots = _generate_time_slots(settings)
    break_slots = {s for s in time_slots if _is_break_slot(s, settings)}

    prev_date = current_date - timedelta(days=1)
    next_date = current_date + timedelta(days=1)
    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return render_template(
        'schedule/day.html',
        current_date=current_date,
        prev_date=prev_date,
        next_date=next_date,
        blocks=enriched,
        time_slots=time_slots,
        break_slots=break_slots,
        settings=settings,
        queue_tasks=queue_tasks,
    )


@schedule_bp.route('/week')
def week_view():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    week_start = current_date - timedelta(days=current_date.weekday())
    week_end = week_start + timedelta(days=6)
    week_days = [week_start + timedelta(days=i) for i in range(7)]

    blocks = schedule_repo.get_by_date_range(
        week_start.isoformat(), week_end.isoformat(),
    )
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    blocks_by_date = {}
    for b in enriched:
        blocks_by_date.setdefault(b['date'], []).append(b)
    for day_key in blocks_by_date:
        blocks_by_date[day_key] = _compute_overlap_layout(blocks_by_date[day_key])

    time_slots = _generate_time_slots(settings)
    break_slots = {s for s in time_slots if _is_break_slot(s, settings)}

    prev_week = current_date - timedelta(weeks=1)
    next_week = current_date + timedelta(weeks=1)
    day_names = ['월', '화', '수', '목', '금', '토', '일']
    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return render_template(
        'schedule/week.html',
        current_date=current_date,
        week_start=week_start,
        week_end=week_end,
        week_days=week_days,
        day_names=day_names,
        prev_date=prev_week,
        next_date=next_week,
        blocks_by_date=blocks_by_date,
        time_slots=time_slots,
        break_slots=break_slots,
        settings=settings,
        today=date.today(),
        queue_tasks=queue_tasks,
    )


@schedule_bp.route('/month')
def month_view():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    year = current_date.year
    month = current_date.month

    cal = calendar.Calendar(firstweekday=0)
    month_days = cal.monthdayscalendar(year, month)

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    blocks = schedule_repo.get_by_date_range(
        first_day.isoformat(), last_day.isoformat(),
    )
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    blocks_by_date = {}
    for b in enriched:
        blocks_by_date.setdefault(b['date'], []).append(b)

    weeks = []
    for week in month_days:
        week_data = []
        for day_num in week:
            if day_num == 0:
                week_data.append(None)
            else:
                d = date(year, month, day_num)
                day_blocks = blocks_by_date.get(d.isoformat(), [])
                week_data.append({
                    'date': d,
                    'day': day_num,
                    'blocks': day_blocks,
                })
        weeks.append(week_data)

    if month == 1:
        prev_month_date = date(year - 1, 12, 1)
    else:
        prev_month_date = date(year, month - 1, 1)
    if month == 12:
        next_month_date = date(year + 1, 1, 1)
    else:
        next_month_date = date(year, month + 1, 1)

    day_names = ['월', '화', '수', '목', '금', '토', '일']
    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return render_template(
        'schedule/month.html',
        current_date=current_date,
        year=year,
        month=month,
        weeks=weeks,
        day_names=day_names,
        prev_date=prev_month_date,
        next_date=next_month_date,
        today=date.today(),
        settings=settings,
        queue_tasks=queue_tasks,
    )


# ---------------------------------------------------------------------------
# View data API endpoints
# ---------------------------------------------------------------------------

@schedule_bp.route('/api/day')
def api_day_data():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    blocks = schedule_repo.get_by_date(current_date.isoformat())
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    time_slots = _generate_time_slots(settings)
    break_slots = [s for s in time_slots if _is_break_slot(s, settings)]

    prev_date = current_date - timedelta(days=1)
    next_date = current_date + timedelta(days=1)

    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return jsonify({
        'current_date': current_date.isoformat(),
        'prev_date': prev_date.isoformat(),
        'next_date': next_date.isoformat(),
        'blocks': enriched,
        'time_slots': time_slots,
        'break_slots': break_slots,
        'settings': settings,
        'queue_tasks': queue_tasks,
    })


@schedule_bp.route('/api/week')
def api_week_data():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    week_start = current_date - timedelta(days=current_date.weekday())
    week_end = week_start + timedelta(days=6)
    week_days = [(week_start + timedelta(days=i)).isoformat() for i in range(7)]

    blocks = schedule_repo.get_by_date_range(week_start.isoformat(), week_end.isoformat())
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    blocks_by_date = {}
    for b in enriched:
        blocks_by_date.setdefault(b['date'], []).append(b)

    time_slots = _generate_time_slots(settings)
    break_slots = [s for s in time_slots if _is_break_slot(s, settings)]

    prev_date = (current_date - timedelta(weeks=1)).isoformat()
    next_date = (current_date + timedelta(weeks=1)).isoformat()

    day_names = ['월', '화', '수', '목', '금', '토', '일']
    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return jsonify({
        'current_date': current_date.isoformat(),
        'week_start': week_start.isoformat(),
        'week_end': week_end.isoformat(),
        'week_days': week_days,
        'day_names': day_names,
        'prev_date': prev_date,
        'next_date': next_date,
        'blocks_by_date': blocks_by_date,
        'time_slots': time_slots,
        'break_slots': break_slots,
        'settings': settings,
        'today': date.today().isoformat(),
        'queue_tasks': queue_tasks,
    })


@schedule_bp.route('/api/month')
def api_month_data():
    current_date = _parse_date(request.args.get('date'))
    settings = settings_repo.get()
    users_map, tasks_map, categories_map = _build_maps()

    year = current_date.year
    month = current_date.month

    cal = calendar.Calendar(firstweekday=0)
    month_days = cal.monthdayscalendar(year, month)

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    blocks = schedule_repo.get_by_date_range(first_day.isoformat(), last_day.isoformat())
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    blocks_by_date = {}
    for b in enriched:
        blocks_by_date.setdefault(b['date'], []).append(b)

    weeks = []
    for week in month_days:
        week_data = []
        for day_num in week:
            if day_num == 0:
                week_data.append(None)
            else:
                d = date(year, month, day_num)
                day_blocks = blocks_by_date.get(d.isoformat(), [])
                week_data.append({
                    'date': d.isoformat(),
                    'day': day_num,
                    'blocks': day_blocks,
                })
        weeks.append(week_data)

    if month == 1:
        prev_month_date = date(year - 1, 12, 1)
    else:
        prev_month_date = date(year, month - 1, 1)
    if month == 12:
        next_month_date = date(year + 1, 1, 1)
    else:
        next_month_date = date(year, month + 1, 1)

    day_names = ['월', '화', '수', '목', '금', '토', '일']
    queue_tasks = _get_queue_tasks(users_map, categories_map)

    return jsonify({
        'current_date': current_date.isoformat(),
        'year': year,
        'month': month,
        'weeks': weeks,
        'day_names': day_names,
        'prev_date': prev_month_date.isoformat(),
        'next_date': next_month_date.isoformat(),
        'today': date.today().isoformat(),
        'settings': settings,
        'queue_tasks': queue_tasks,
    })


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------

@schedule_bp.route('/api/blocks', methods=['POST'])
def api_create_block():
    data = request.get_json()
    if not data:
        return jsonify({'error': '요청 데이터가 없습니다.'}), 400

    required = ['task_id', 'date', 'start_time', 'end_time']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field}은(는) 필수 항목입니다.'}), 400

    # If no assignee_id, use the task's assignee
    assignee_id = data.get('assignee_id', '')
    if not assignee_id:
        task = task_repo.get_by_id(data['task_id'])
        if task:
            assignee_id = task.get('assignee_id', '')

    # Adjust end_time to skip breaks/lunch
    settings = settings_repo.get()
    adjusted_end = _adjust_end_for_breaks(
        data['start_time'], data['end_time'], settings,
    )

    # Check overlap
    overlap = _check_overlap(assignee_id, data['date'], data['start_time'], adjusted_end)
    if overlap:
        return jsonify({'error': '해당 시간에 이미 다른 업무가 배치되어 있습니다.'}), 409

    block = schedule_repo.create(
        task_id=data['task_id'],
        assignee_id=assignee_id,
        date=data['date'],
        start_time=data['start_time'],
        end_time=adjusted_end,
        is_draft=data.get('is_draft', False),
        is_locked=data.get('is_locked', False),
        origin=data.get('origin', 'manual'),
    )
    return jsonify(block), 201


@schedule_bp.route('/api/blocks/<block_id>', methods=['PUT'])
def api_update_block(block_id):
    block = schedule_repo.get_by_id(block_id)
    if not block:
        return jsonify({'error': '블록을 찾을 수 없습니다.'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'error': '요청 데이터가 없습니다.'}), 400

    allowed = ['date', 'start_time', 'end_time', 'is_draft', 'is_locked']
    updates = {k: v for k, v in data.items() if k in allowed}
    is_resize = data.get('resize', False)

    # Adjust end_time: preserve original work duration at new position
    # Skip this for resize operations — the user explicitly chose the new size
    if 'start_time' in updates and 'end_time' in updates and not is_resize:
        settings = settings_repo.get()
        # Get actual work minutes of original block (excluding breaks)
        work_mins = _work_minutes_in_range(
            block['start_time'], block['end_time'], settings,
        )
        # Compute raw end = new_start + work_duration, then adjust for breaks
        raw_end = _minutes_to_time(
            _time_to_minutes(updates['start_time']) + work_mins,
        )
        updates['end_time'] = _adjust_end_for_breaks(
            updates['start_time'], raw_end, settings,
        )

    # Check overlap
    check_date = updates.get('date', block['date'])
    check_start = updates.get('start_time', block['start_time'])
    check_end = updates.get('end_time', block['end_time'])
    overlap = _check_overlap(block['assignee_id'], check_date, check_start, check_end, exclude_block_id=block_id)
    if overlap:
        return jsonify({'error': '해당 시간에 이미 다른 업무가 배치되어 있습니다.'}), 409

    updated = schedule_repo.update(block_id, **updates)

    # On resize, lock in the task's remaining_hours to match total scheduled time
    if is_resize:
        task_id = block.get('task_id', '')
        if task_id:
            all_blocks = schedule_repo.get_all()
            total_scheduled_min = 0
            for b in all_blocks:
                if b['task_id'] == task_id:
                    total_scheduled_min += _time_to_minutes(b['end_time']) - _time_to_minutes(b['start_time'])
            total_scheduled_hours = round(total_scheduled_min / 60.0, 2)
            task = task_repo.get_by_id(task_id)
            if task and task.get('remaining_hours', 0) != total_scheduled_hours:
                task_repo.patch(task_id, remaining_hours=total_scheduled_hours)

    return jsonify(updated)


@schedule_bp.route('/api/blocks/<block_id>', methods=['DELETE'])
def api_delete_block(block_id):
    block = schedule_repo.get_by_id(block_id)
    if not block:
        return jsonify({'error': '블록을 찾을 수 없습니다.'}), 404

    schedule_repo.delete(block_id)
    return jsonify({'success': True})


@schedule_bp.route('/api/blocks/<block_id>/lock', methods=['PUT'])
def api_toggle_lock(block_id):
    block = schedule_repo.get_by_id(block_id)
    if not block:
        return jsonify({'error': '블록을 찾을 수 없습니다.'}), 404

    new_locked = not block.get('is_locked', False)
    updated = schedule_repo.update(block_id, is_locked=new_locked)
    return jsonify(updated)


@schedule_bp.route('/api/export')
def api_export():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    fmt = request.args.get('format', 'csv')

    if not start_date or not end_date:
        return jsonify({'error': 'start_date와 end_date는 필수입니다.'}), 400

    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': '날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)'}), 400

    users_map, tasks_map, categories_map = _build_maps()
    settings = settings_repo.get()
    blocks = schedule_repo.get_by_date_range(start_date, end_date)
    enriched = _enrich_blocks(
        blocks, users_map, tasks_map, categories_map,
        settings.get('block_color_by', 'assignee'),
    )

    # Sort by date, then start_time
    enriched.sort(key=lambda b: (b.get('date', ''), b.get('start_time', '')))

    headers = ['날짜', '시작시간', '종료시간', '업무명', '담당자', '카테고리', '우선순위', '상태', '잠금']

    def block_to_row(b):
        status = '초안' if b.get('is_draft') else '확정'
        locked = 'Y' if b.get('is_locked') else 'N'
        priority_map = {'high': '높음', 'medium': '보통', 'low': '낮음'}
        return [
            b.get('date', ''),
            b.get('start_time', ''),
            b.get('end_time', ''),
            b.get('task_title', ''),
            b.get('assignee_name', ''),
            b.get('category_name', ''),
            priority_map.get(b.get('priority', ''), b.get('priority', '')),
            status,
            locked,
        ]

    # Try xlsx if requested — calendar layout
    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = '스케줄'

            # -- Group blocks by date --
            blocks_by_date = {}
            for b in enriched:
                d = b.get('date', '')
                blocks_by_date.setdefault(d, []).append(b)

            # -- Build calendar weeks --
            d_start = datetime.strptime(start_date, '%Y-%m-%d').date()
            d_end = datetime.strptime(end_date, '%Y-%m-%d').date()
            # Expand to full weeks (Mon-Sun)
            week_start = d_start - timedelta(days=d_start.weekday())
            week_end = d_end + timedelta(days=6 - d_end.weekday())

            day_names = ['월', '화', '수', '목', '금', '토', '일']

            # Styles
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font_white = Font(bold=True, size=11, color='FFFFFF')
            date_font = Font(bold=True, size=10)
            block_font = Font(size=9)
            today_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
            weekend_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
            top_align = Alignment(vertical='top', wrap_text=True)
            today = date.today()

            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            title_cell = ws.cell(row=1, column=1, value=f'스케줄: {start_date} ~ {end_date}')
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')

            # Column widths
            for c in range(1, 8):
                ws.column_dimensions[get_column_letter(c)].width = 22

            row = 3  # start from row 3

            current = week_start
            while current <= week_end:
                # Day-name header row
                for i in range(7):
                    cell = ws.cell(row=row, column=i + 1, value=day_names[i])
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                row += 1

                # Date row
                for i in range(7):
                    day = current + timedelta(days=i)
                    label = day.strftime('%m/%d')
                    if day < d_start or day > d_end:
                        label = ''
                    cell = ws.cell(row=row, column=i + 1, value=label)
                    cell.font = date_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    if day == today:
                        cell.fill = today_fill
                    elif day.weekday() >= 5:
                        cell.fill = weekend_fill
                row += 1

                # Determine max blocks in this week to size content rows
                max_blocks = 0
                for i in range(7):
                    day = current + timedelta(days=i)
                    day_str = day.isoformat()
                    max_blocks = max(max_blocks, len(blocks_by_date.get(day_str, [])))
                content_rows = max(max_blocks, 1)

                # Fill block content rows
                for r_offset in range(content_rows):
                    for i in range(7):
                        day = current + timedelta(days=i)
                        day_str = day.isoformat()
                        day_blocks = blocks_by_date.get(day_str, [])

                        cell = ws.cell(row=row + r_offset, column=i + 1)
                        cell.border = thin_border
                        cell.alignment = top_align

                        if day.weekday() >= 5:
                            cell.fill = weekend_fill
                        if day == today:
                            cell.fill = today_fill

                        if r_offset < len(day_blocks):
                            b = day_blocks[r_offset]
                            priority_map = {'high': '⬆높음', 'medium': '➡보통', 'low': '⬇낮음'}
                            pri = priority_map.get(b.get('priority', ''), '')
                            lines = [
                                f"{b.get('start_time', '')}~{b.get('end_time', '')}",
                                b.get('task_title', ''),
                                b.get('assignee_name', ''),
                            ]
                            if b.get('category_name'):
                                lines.append(f"[{b['category_name']}]")
                            if pri:
                                lines.append(pri)
                            cell.value = '\n'.join(lines)
                            cell.font = block_font
                            # Tint cell with block color
                            color_hex = (b.get('color') or '#FFFFFF').lstrip('#')
                            if len(color_hex) == 6:
                                # Lighten: blend with white at 30% opacity
                                r_c = int(color_hex[0:2], 16)
                                g_c = int(color_hex[2:4], 16)
                                b_c = int(color_hex[4:6], 16)
                                r_c = int(r_c * 0.3 + 255 * 0.7)
                                g_c = int(g_c * 0.3 + 255 * 0.7)
                                b_c = int(b_c * 0.3 + 255 * 0.7)
                                light = f'{r_c:02X}{g_c:02X}{b_c:02X}'
                                cell.fill = PatternFill(start_color=light, end_color=light, fill_type='solid')

                # Set row heights for content rows
                for r_offset in range(content_rows):
                    ws.row_dimensions[row + r_offset].height = 60

                row += content_rows
                row += 1  # blank separator row between weeks
                current += timedelta(days=7)

            # -- Data sheet for raw data --
            ws2 = wb.create_sheet(title='데이터')
            ws2.append(headers)
            for b in enriched:
                ws2.append(block_to_row(b))
            for col in ws2.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value) if cell.value else ''
                    max_len = max(max_len, len(val))
                ws2.column_dimensions[col[0].column_letter].width = max_len + 4

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f'schedule_{start_date}_{end_date}.xlsx'
            return Response(
                buf.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{filename}"'},
            )
        except ImportError:
            # Fall back to CSV
            fmt = 'csv'

    # CSV output
    buf = io.StringIO()
    # BOM for Excel compatibility with Korean
    buf.write('\ufeff')
    writer = csv.writer(buf)
    writer.writerow(headers)
    for b in enriched:
        writer.writerow(block_to_row(b))
    filename = f'schedule_{start_date}_{end_date}.csv'
    return Response(
        buf.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@schedule_bp.route('/api/draft/generate', methods=['POST'])
def api_draft_generate():
    from app.services.scheduler import generate_draft_schedule
    result = generate_draft_schedule()
    return jsonify({
        'placed_count': len(result['placed']),
        'unplaced': [
            {
                'task_id': u['task']['id'],
                'task_title': u['task']['title'],
                'remaining_hours': u['remaining_unscheduled_hours'],
            }
            for u in result['unplaced']
        ],
    })


@schedule_bp.route('/api/draft/approve', methods=['POST'])
def api_draft_approve():
    from app.services.scheduler import approve_drafts
    approve_drafts()
    return jsonify({'success': True})


@schedule_bp.route('/api/draft/discard', methods=['POST'])
def api_draft_discard():
    from app.services.scheduler import discard_drafts
    discard_drafts()
    return jsonify({'success': True})
