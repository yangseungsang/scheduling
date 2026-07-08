"""스케줄 내보내기(export) 서비스 모듈.

보강된 스케줄 블록 데이터를 CSV 또는 XLSX(엑셀) 형식으로 내보내는
기능을 제공한다. 엑셀 내보내기 시 달력 레이아웃 시트와 데이터 시트를
함께 생성한다.
"""

import csv
import io
import zipfile
from datetime import date, datetime, timedelta
from xml.sax.saxutils import escape


# 내보내기 열 헤더 (기존 Excel 양식)
HEADERS = ['날짜', '문서명']
PRACTITIONER_HEADERS = ['날짜', '장소', '시작', '종료', '절차서', '시험 식별자']


def _block_to_row(b):
    """블록 딕셔너리를 내보내기용 행(row) 데이터로 변환한다."""
    name = b.get('display_name') or b.get('doc_name', '') or b.get('task_title', '')
    if b.get('is_split'):
        name += (
            ' ('
            + str(b.get('block_identifier_count', '?'))
            + '/'
            + str(b.get('total_identifier_count', '?'))
            + ')'
        )
    return [b.get('date', ''), name]


def _block_label(b):
    """블록의 표시 라벨을 생성한다 (달력 시트용)."""
    name = b.get('display_name') or b.get('doc_name', '') or b.get('task_title', '')
    if b.get('is_split'):
        name += (
            ' ('
            + str(b.get('block_identifier_count', '?'))
            + '/'
            + str(b.get('total_identifier_count', '?'))
            + ')'
        )
    return name


def _identifier_label(identifier):
    if isinstance(identifier, dict):
        identifier_id = identifier.get('id', '')
        identifier_name = identifier.get('name', '')
        if identifier_id and identifier_name:
            return f'{identifier_id} - {identifier_name}'
        return identifier_id or identifier_name
    return str(identifier) if identifier is not None else ''


def _block_identifiers(b):
    identifiers = b.get('identifiers') or []
    selected_ids = b.get('identifier_ids')
    if selected_ids is None:
        return identifiers

    selected_id_set = set(selected_ids)
    return [
        identifier
        for identifier in identifiers
        if (
            identifier.get('id') if isinstance(identifier, dict) else identifier
        )
        in selected_id_set
    ]


def _practitioner_rows(enriched_blocks):
    rows = [PRACTITIONER_HEADERS]
    sorted_blocks = sorted(
        enriched_blocks,
        key=lambda b: (
            b.get('date', ''),
            b.get('location_name', ''),
            b.get('start_time', ''),
            b.get('end_time', ''),
            _block_label(b),
        ),
    )
    for b in sorted_blocks:
        identifiers = _block_identifiers(b)
        if not identifiers:
            identifiers = ['']
        for identifier in identifiers:
            rows.append(
                [
                    b.get('date', ''),
                    b.get('location_name', ''),
                    b.get('start_time', ''),
                    b.get('end_time', ''),
                    _block_label(b),
                    _identifier_label(identifier),
                ]
            )
    return rows


def _sheet_xml(rows, styled_rows=None):
    styled_rows = styled_rows or {}

    def col_name(index):
        name = ''
        while index:
            index, rem = divmod(index - 1, 26)
            name = chr(65 + rem) + name
        return name

    xml_rows = []
    for r_idx, row in enumerate(rows, 1):
        cells = []
        style = styled_rows.get(r_idx, 0)
        for c_idx, value in enumerate(row, 1):
            ref = f'{col_name(c_idx)}{r_idx}'
            text = escape(str(value) if value is not None else '')
            style_attr = f' s="{style}"' if style else ''
            cells.append(
                f'<c r="{ref}" t="inlineStr"{style_attr}><is><t>{text}</t></is></c>'
            )
        xml_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
        '<sheetData>' + ''.join(xml_rows) + '</sheetData></worksheet>'
    )


def _export_xlsx_stdlib(enriched_blocks, start_date, end_date, version_name=''):
    """openpyxl이 없는 환경에서도 기본 서식이 있는 XLSX를 생성한다."""
    d_start = datetime.strptime(start_date, '%Y-%m-%d').date()
    d_end = datetime.strptime(end_date, '%Y-%m-%d').date()
    blocks_by_date = {}
    for b in enriched_blocks:
        blocks_by_date.setdefault(b.get('date', ''), []).append(b)

    calendar_rows = [
        [
            f'스케줄: {start_date} ~ {end_date}'
            + (f' / {version_name}' if version_name else '')
        ]
    ]
    day_names = ['월', '화', '수', '목', '금', '토', '일']
    calendar_rows.append(day_names)
    current = d_start
    while current <= d_end:
        week_days = [current + timedelta(days=i) for i in range(7)]
        calendar_rows.append(
            [
                day.strftime('%m/%d') if d_start <= day <= d_end else ''
                for day in week_days
            ]
        )
        max_blocks = max(
            [len(blocks_by_date.get(day.isoformat(), [])) for day in week_days] + [1]
        )
        for block_idx in range(max_blocks):
            calendar_rows.append(
                [
                    _block_label(blocks_by_date.get(day.isoformat(), [])[block_idx])
                    if block_idx < len(blocks_by_date.get(day.isoformat(), []))
                    else ''
                    for day in week_days
                ]
            )
        calendar_rows.append([''] * 7)
        current += timedelta(days=7)

    data_rows = [HEADERS] + [_block_to_row(b) for b in enriched_blocks]
    practitioner_rows = _practitioner_rows(enriched_blocks)
    calendar_styles = {1: 1, 2: 1}
    data_styles = {1: 1}
    practitioner_styles = {1: 1}

    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF4472C4"/><bgColor indexed="64"/></patternFill></fill></fills>
  <borders count="2"><border/><border><left style="thin"/><right style="thin"/><top style="thin"/><bottom style="thin"/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf><xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf></cellXfs>
</styleSheet>"""

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            '[Content_Types].xml',
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet3.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>""",
        )
        z.writestr(
            '_rels/.rels',
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>""",
        )
        z.writestr(
            'xl/workbook.xml',
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="스케줄" sheetId="1" r:id="rId1"/><sheet name="데이터" sheetId="2" r:id="rId2"/><sheet name="실무자용" sheetId="3" r:id="rId3"/></sheets></workbook>""",
        )
        z.writestr(
            'xl/_rels/workbook.xml.rels',
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet3.xml"/><Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>""",
        )
        z.writestr('xl/styles.xml', styles_xml)
        z.writestr(
            'xl/worksheets/sheet1.xml', _sheet_xml(calendar_rows, calendar_styles)
        )
        z.writestr('xl/worksheets/sheet2.xml', _sheet_xml(data_rows, data_styles))
        z.writestr(
            'xl/worksheets/sheet3.xml',
            _sheet_xml(practitioner_rows, practitioner_styles),
        )
    return buf.getvalue()


def export_csv(enriched_blocks):
    """보강된 블록 목록을 CSV 문자열로 변환한다.

    한국어 엑셀 호환성을 위해 BOM(Byte Order Mark)을 앞에 추가한다.

    Args:
        enriched_blocks: 보강된 스케줄 블록 딕셔너리 목록.

    Returns:
        str: BOM이 포함된 CSV 문자열.
    """
    buf = io.StringIO()
    buf.write('\ufeff')  # UTF-8 BOM: 엑셀에서 한글 깨짐 방지
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for b in enriched_blocks:
        writer.writerow(_block_to_row(b))
    return buf.getvalue()


def export_xlsx(enriched_blocks, start_date, end_date, version_name=''):
    """보강된 블록 목록을 달력 형태의 XLSX 파일로 변환한다.

    첫 번째 시트('스케줄')에는 주간 달력 레이아웃으로 각 날짜에
    해당하는 장절명을 표시하고, 두 번째 시트('데이터')에는
    원본 데이터를 목록 형태로 나열한다.

    Args:
        enriched_blocks: 보강된 스케줄 블록 딕셔너리 목록.
        start_date: 내보내기 시작 날짜 ('YYYY-MM-DD').
        end_date: 내보내기 종료 날짜 ('YYYY-MM-DD').
        version_name: 소프트웨어 버전명 (선택).

    Returns:
        bytes: XLSX 파일 바이너리 데이터.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _export_xlsx_stdlib(enriched_blocks, start_date, end_date, version_name)

    wb = Workbook()
    ws = wb.active
    ws.title = '스케줄'

    # 블록을 날짜별로 그룹핑
    blocks_by_date = {}
    for b in enriched_blocks:
        blocks_by_date.setdefault(b.get('date', ''), []).append(b)

    # 달력 주간 범위 계산: 시작일이 속한 주의 월요일 ~ 종료일이 속한 주의 일요일
    d_start = datetime.strptime(start_date, '%Y-%m-%d').date()
    d_end = datetime.strptime(end_date, '%Y-%m-%d').date()
    week_start = d_start - timedelta(days=d_start.weekday())  # 월요일로 맞춤
    week_end = d_end + timedelta(days=6 - d_end.weekday())

    day_names = ['월', '화', '수', '목', '금', '토', '일']

    # 엑셀 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(
        start_color='4472C4', end_color='4472C4', fill_type='solid'
    )
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    date_font = Font(bold=True, size=10)
    block_font = Font(size=9)
    today_fill = PatternFill(
        start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'
    )
    weekend_fill = PatternFill(
        start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'
    )
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    top_align = Alignment(vertical='top', wrap_text=True)
    today = date.today()

    # 1행: 제목 (7열 병합)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_text = f'스케줄: {start_date} ~ {end_date}'
    if version_name:
        title_text = f'[{version_name}] {title_text}'
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # 열 너비 설정 (7일 = 7열)
    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # 주간 반복: 요일 행 + 날짜 행 + 블록 행 (월~일)
    row = 3
    current = week_start
    while current <= week_end:
        # 요일 헤더 행
        for i in range(7):
            cell = ws.cell(row=row, column=i + 1, value=day_names[i])
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        # 날짜 행
        for i in range(7):
            day = current + timedelta(days=i)
            label = day.strftime('%m/%d') if d_start <= day <= d_end else ''
            cell = ws.cell(row=row, column=i + 1, value=label)
            cell.font = date_font
            cell.alignment = center_align
            cell.border = thin_border
            if day == today:
                cell.fill = today_fill
            elif day.weekday() >= 5:
                cell.fill = weekend_fill
        row += 1

        max_blocks = 0
        for i in range(7):
            day = current + timedelta(days=i)
            max_blocks = max(max_blocks, len(blocks_by_date.get(day.isoformat(), [])))
        content_rows = max(max_blocks, 1)

        for r_offset in range(content_rows):
            for i in range(7):
                day = current + timedelta(days=i)
                day_blocks = blocks_by_date.get(day.isoformat(), [])

                cell = ws.cell(row=row + r_offset, column=i + 1)
                cell.border = thin_border
                cell.alignment = top_align

                if day.weekday() >= 5:
                    cell.fill = weekend_fill
                if day == today:
                    cell.fill = today_fill

                if r_offset < len(day_blocks):
                    block = day_blocks[r_offset]
                    cell.value = _block_label(block)
                    cell.font = block_font
                    color_hex = (block.get('color') or '#FFFFFF').lstrip('#')
                    if len(color_hex) == 6:
                        r_c = int(int(color_hex[0:2], 16) * 0.3 + 255 * 0.7)
                        g_c = int(int(color_hex[2:4], 16) * 0.3 + 255 * 0.7)
                        b_c = int(int(color_hex[4:6], 16) * 0.3 + 255 * 0.7)
                        light = f'{r_c:02X}{g_c:02X}{b_c:02X}'
                        cell.fill = PatternFill(
                            start_color=light, end_color=light, fill_type='solid'
                        )

        for r_offset in range(content_rows):
            ws.row_dimensions[row + r_offset].height = 60

        row += content_rows
        row += 1

        # 다음 주로 이동
        current += timedelta(days=7)

    # 두 번째 시트: 원본 데이터 목록
    ws2 = wb.create_sheet(title='데이터')
    ws2.append(HEADERS)
    for b in enriched_blocks:
        ws2.append(_block_to_row(b))
    # 열 너비를 내용에 맞게 자동 조정
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ''
            max_len = max(max_len, len(val))
        ws2.column_dimensions[col[0].column_letter].width = max_len + 4

    # 세 번째 시트: 실무자용 날짜/장소/식별자 목록
    ws3 = wb.create_sheet(title='실무자용')
    for row_values in _practitioner_rows(enriched_blocks):
        ws3.append(row_values)
    for cell in ws3[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    for row_cells in ws3.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = top_align
            cell.border = thin_border
    ws3.freeze_panes = 'A2'
    for col in ws3.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ''
            max_len = max(max_len, len(val))
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # 워크북을 바이트 버퍼에 저장하여 반환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
